import glob
import csv
import os
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# # Settings
# input_folder = "whisperx_output"
# output_xlsx = os.path.join(input_folder, "full_transcript.xlsx")
# output_txt = os.path.join(input_folder, "full_transcript.txt")
# CONFIG_FILE = "config.json"

# Color palette for speakers
SPEAKER_COLORS = [
    "FFFFCC", "CCFFCC", "CCE5FF", "FFCCCC", "FFCC99",
    "CCFFFF", "E0CCFF", "FFCCE5", "D9D9D9", "CCE5CC"
]

# Load config.json
def load_config(config_path):
    if not os.path.exists(config_path):
        print(f"❌ Config file not found: {config_path}")
        return {}
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)

# Build character -> player mapping from config
def build_character_to_player_map(config_data):
    mapping = {}
    campaigns = config_data.get("campaigns", {})
    for campaign, characters in campaigns.items():
        for character, player in characters.items():
            mapping[character.lower()] = player  # Lowercase for safer matching
    return mapping

def load_tsv_entries(tsv_path):
    entries = []
    with open(tsv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            try:
                entries.append({
                    "start": float(row["Start"]),
                    "end": float(row["End"]),
                    "speaker": row["Speaker"],
                    "text": row["Text"]
                })
            except (KeyError, ValueError) as e:
                print(f"⚠️ Skipping bad row in {tsv_path}: {row}")
    return entries

def main():
    # Use batch file settings if defined, otherwise use the below settings for output file naming
    if len(sys.argv) < 2:
        print("❌ Usage: combine_tsvs_with_colors.py [input_folder] [output_xlsx] [output_txt]")
        return

    input_folder = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) > 2 else os.path.join(input_folder, "full_transcript.xlsx")
    output_txt = sys.argv[3] if len(sys.argv) > 3 else os.path.join(input_folder, "full_transcript.txt")
    
    # Load config and character-player mapping
    # config_data = load_config(CONFIG_FILE)
    config_data = load_config("config.json")
    character_to_player = build_character_to_player_map(config_data)

    # Find all TSVs
    tsv_files = glob.glob(os.path.join(input_folder, "**", "*.tsv"), recursive=True)
    tsv_files = [f for f in tsv_files if "full_transcript" not in os.path.basename(f).lower()]
    print(f"📁 Found {len(tsv_files)} TSV files")
    
    entries = []
    for tsv_file in tsv_files:
        entries.extend(load_tsv_entries(tsv_file))        
        # if "full_transcript" in os.path.basename(tsv_file).lower():
            # continue
        # entries.extend(load_tsv_entries(tsv_file))

    if not entries:
        print("❌ No valid entries found. Exiting.")
        return

    # Sort by start time
    entries.sort(key=lambda x: x["start"])
    
    # Get list of players
    players = sorted(set(
        character_to_player.get(entry["speaker"].lower(), entry["speaker"])
        for entry in entries
    ))

    player_color_map = {
        player: SPEAKER_COLORS[i % len(SPEAKER_COLORS)]
        for i, player in enumerate(players)
    }

    # Create XLSX
    wb = Workbook()
    ws = wb.active
    ws.append(["Start", "End", "Player", "Text"])

    for entry in entries:
        player = character_to_player.get(entry["speaker"].lower(), entry["speaker"])
        color = player_color_map.get(player, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        row = [entry["start"], entry["end"], player, entry["text"]]
        ws.append(row)
        for col_idx in range(1, 5):
            ws.cell(row=ws.max_row, column=col_idx).fill = fill

    wb.save(output_xlsx)
    print(f"✅ Created XLSX: {output_xlsx}")

    # Create TXT
    with open(output_txt, "w", encoding="utf-8") as f:
        for entry in entries:
            player = character_to_player.get(entry["speaker"].lower(), entry["speaker"])
            f.write(f"[{player}] {entry['text']}\n")

    print(f"✅ Created TXT: {output_txt}")

if __name__ == "__main__":
    main()



# import glob
# import csv
# import os
# import json
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill

# # Settings
# input_folder = "whisperx_output"
# output_xlsx = os.path.join(input_folder, "full_transcript.xlsx")
# output_txt = os.path.join(input_folder, "full_transcript.txt")
# CONFIG_FILE = "config.json"

# # Color palette for speakers
# SPEAKER_COLORS = [
    # "FFFFCC", "CCFFCC", "CCE5FF", "FFCCCC", "FFCC99",
    # "CCFFFF", "E0CCFF", "FFCCE5", "D9D9D9", "CCE5CC"
# ]

# # Load config.json
# def load_config(config_path):
    # if not os.path.exists(config_path):
        # print(f"❌ Config file not found: {config_path}")
        # return {}
    # with open(config_path, "r", encoding="utf-8") as f:
        # return json.load(f)

# # Build character -> player mapping from config
# def build_character_to_player_map(config_data):
    # mapping = {}
    # campaigns = config_data.get("campaigns", {})
    # for campaign, characters in campaigns.items():
        # for character, player in characters.items():
            # mapping[character.lower()] = player  # Lowercase for safer matching
    # return mapping

# def load_tsv_entries(tsv_path):
    # entries = []
    # with open(tsv_path, "r", encoding="utf-8") as f:
        # reader = csv.DictReader(f, delimiter="\t")
        # for row in reader:
            # try:
                # entries.append({
                    # "start": float(row["Start"]),
                    # "end": float(row["End"]),
                    # "speaker": row["Speaker"],
                    # "text": row["Text"]
                # })
            # except (KeyError, ValueError) as e:
                # print(f"⚠️ Skipping bad row in {tsv_path}: {row}")
    # return entries

# def main():
    # # Load config and character-player mapping
    # config_data = load_config(CONFIG_FILE)
    # character_to_player = build_character_to_player_map(config_data)

    # # Find all TSVs
    # tsv_files = glob.glob(os.path.join(input_folder, "**", "*.tsv"), recursive=True)

    # entries = []
    # for tsv_file in tsv_files:
        # if "full_transcript" in os.path.basename(tsv_file).lower():
            # continue
        # entries.extend(load_tsv_entries(tsv_file))

    # if not entries:
        # print("❌ No valid entries found. Exiting.")
        # return
    
    # # Recursively find all .tsv files
    # tsv_files = glob.glob(os.path.join(input_folder, "**", "*.tsv"), recursive=True)

    # entries = []
    # for tsv_file in tsv_files:
        # if "full_transcript" in os.path.basename(tsv_file).lower():
            # continue  # Skip any old full transcripts
        # entries.extend(load_tsv_entries(tsv_file))

    # if not entries:
        # print("❌ No valid entries found. Exiting.")
        # return

    # Sort by start time
    # entries.sort(key=lambda x: x["start"])
    
    # # Get list of players
    # players = sorted(set(
        # character_to_player.get(entry["speaker"].lower(), entry["speaker"])
        # for entry in entries
    # ))

    # player_color_map = {
        # player: PLAYER_COLORS[i % len(PLAYER_COLORS)]
        # for i, player in enumerate(players)
    # }

    # # Assign colors to speakers
    # # speakers = sorted(set(entry["speaker"] for entry in entries))
    # # speaker_color_map = {
        # # speaker: SPEAKER_COLORS[i % len(SPEAKER_COLORS)]
        # # for i, speaker in enumerate(speakers)
    # #}

    # # Create XLSX
    # wb = Workbook()
    # ws = wb.active
    # ws.append(["Start", "End", "Speaker", "Text"])

    # for entry in entries:
        # # Map speaker to player name
        # player = character_to_player.get(entry["speaker"].lower(), entry["speaker"])
        # color = player_color_map.get(player, "FFFFFF")
        # fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # row = [entry["start"], entry["end"], entry["speaker"], entry["text"]]
        # ws.append(row)
        # for col_idx in range(1, 5):
            # ws.cell(row=ws.max_row, column=col_idx).fill = fill

    # # for entry in entries:
        # # color = speaker_color_map.get(entry["speaker"], "FFFFFF")
        # # fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        # # row = [entry["start"], entry["end"], entry["speaker"], entry["text"]]
        # # ws.append(row)
        # # for col_idx in range(1, 5):
            # # ws.cell(row=ws.max_row, column=col_idx).fill = fill

    # wb.save(output_xlsx)
    # print(f"✅ Created XLSX: {output_xlsx}")

    # # Create TXT
    # with open(output_txt, "w", encoding="utf-8") as f:
        # for entry in entries:
            # f.write(f"[{entry['speaker']}] {entry['text']}\n")

    # print(f"✅ Created TXT: {output_txt}")

# if __name__ == "__main__":
    # main()

# Old Code ------------
# import glob
# import csv
# import os
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill

# # Settings
# input_folder = "whisperx_output"
# output_xlsx = os.path.join(input_folder, "full_transcript.xlsx")
# output_txt = os.path.join(input_folder, "full_transcript.txt")

# # Define a color palette for speakers (feel free to add more colors!)
# SPEAKER_COLORS = [
    # "FFFFCC", "CCFFCC", "CCE5FF", "FFCCCC", "FFCC99",
    # "CCFFFF", "E0CCFF", "FFCCE5", "D9D9D9", "CCE5CC"
# ]

# # Load all .tsv files
# tsv_files = glob.glob(os.path.join(input_folder, "**", "*.tsv"), recursive=True)

# entries = []

# # Read each .tsv
# for tsv_file in tsv_files:
    # if os.path.basename(tsv_file).lower().startswith("full_"):
        # continue  # skip already combined file
    # with open(tsv_file, "r", encoding="utf-8") as f:
        # reader = csv.DictReader(f, delimiter="\t")
        # for row in reader:
            # entries.append({
                # "start": float(row["Start"]),
                # "end": float(row["End"]),
                # "speaker": row["Speaker"],
                # "text": row["Text"]
            # })

# # Sort by start time
# entries.sort(key=lambda x: x["start"])

# # Collect speakers and assign colors
# speakers = sorted(list(set(entry["speaker"] for entry in entries)))
# speaker_color_map = {
    # speaker: SPEAKER_COLORS[i % len(SPEAKER_COLORS)]
    # for i, speaker in enumerate(speakers)
# }

# # Create XLSX
# wb = Workbook()
# ws = wb.active
# ws.append(["Start", "End", "Speaker", "Text"])

# for entry in entries:
    # color = speaker_color_map.get(entry["speaker"], "FFFFFF")
    # fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    # row = [entry["start"], entry["end"], entry["speaker"], entry["text"]]
    # ws.append(row)
    # for col_idx in range(1, 5):
        # ws.cell(row=ws.max_row, column=col_idx).fill = fill

# wb.save(output_xlsx)
# print(f"✅ Created XLSX: {output_xlsx}")

# # Create TXT
# with open(output_txt, "w", encoding="utf-8") as f:
    # for entry in entries:
        # f.write(f"[{entry['speaker']}] {entry['text']}\n")

# print(f"✅ Created TXT: {output_txt}")
