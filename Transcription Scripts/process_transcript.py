# === Imports and Configuration ===
# These libraries handle file I/O, text cleaning, spell checking, fuzzy matching, logging, and tabular data.
# SpellChecker flags unknown words.
# Levenshtein and jellyfish calculate string similarity (for "galor" ≈ "galore").
# =================================

import sys
import os
import json
import pandas as pd
import difflib
import re
from spellchecker import SpellChecker
import Levenshtein
import jellyfish
from collections import defaultdict, Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Initial Configuration ===
# Defines output file names and suffixes to strip when checking unknown words (e.g., "Tinkler's" → "Tinkler").
# =============================

ANOMALY_LOG_CSV = "anomalies_log.csv"
SUGGESTIONS_FILE = "glossary_suggestions.json"
GLOSSARY_FILE = "glossary_config.json"
SUFFIXES = ("'s", "'ll", "'d", "'re", "'ve", "'m", "’s", "’ll", "’d", "’re", "’ve", "’m")
SPEAKER_COLORS = [
    "FFFFCC", "CCFFCC", "CCE5FF", "FFCCCC", "FFCC99",
    "CCFFFF", "E0CCFF", "FFCCE5", "D9D9D9", "CCE5CC"
]

# === Glossary Loader ===
# Reads glossary_config.json and extracts two sections:
    # "Replace" → dictionary of correct term → [list of incorrect terms]
    # "Ignore" → set of words to skip during anomaly checks (e.g., fantasy names, homebrew terms)
# =======================

def load_glossary(file_path):
    if not os.path.exists(file_path):
        return {}, set()

    with open(file_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Ensure only the nested dictionary is returned
    replace_dict = data.get("Replace", {})
    if not isinstance(replace_dict, dict):
        raise ValueError("🚨 The 'Replace' section in your glossary_config.json must be a dictionary.")

    ignore_words = set(word.lower() for word in data.get("Ignore", []))
    return replace_dict, ignore_words


# === Suffix Stripper ===
# Strips common possessive or contraction suffixes from words.
# "Traveller's" → "Traveller"
# =======================
    
def strip_suffix(word):
    for suf in SUFFIXES:
        if word.lower().endswith(suf):
            return word[:-len(suf)]
    return word

# === Line Combiner ===
# Merges lines by the same speaker while preserving:
    # Start time from first line
    # End time from last line
    # All intermediate text
# Output includes: Player, Adjusted Line, Start, End.
# ======================

def combine_lines(df):
    combined = []
    current_speaker = None
    current_lines = []
    start_time = None
    end_time = None

    for _, row in df.iterrows():
        speaker = row["Player"]
        line = str(row["Text"])
        if speaker == current_speaker:
            current_lines.append(line)
            end_time = row["End"]
        else:
            if current_speaker is not None:
                combined.append({
                    "Player": current_speaker,
                    "Adjusted Line": " ".join(current_lines),
                    "Start": start_time,
                    "End": end_time
                })
            current_speaker = speaker
            current_lines = [line]
            start_time = row["Start"]
            end_time = row["End"]

    if current_speaker is not None:
        combined.append({
            "Player": current_speaker,
            "Adjusted Line": " ".join(current_lines),
            "Start": start_time,
            "End": end_time
        })

    return pd.DataFrame(combined)


# === Apply Glossary Corrections ===
# Replaces each incorrect term (case-insensitive) with the correct one from the glossary.
# Logs changes in format: old_word → correct_word.
# ==================================

def apply_corrections(text, replacements, applied_log):
    summary = []

    for correct, wrongs in replacements.items():
        # Skip if the key is likely malformed (e.g., "Replace" accidentally inside)
        if not isinstance(wrongs, list) or correct.lower() in {"replace", "ignore"}:
            continue

        for wrong in wrongs:
            pattern = r"\b{}\b".format(re.escape(wrong))
            matches = list(re.finditer(pattern, text, flags=re.IGNORECASE))
            for match in matches:
                summary.append(f"{match.group(0)} → {correct}")
            text = re.sub(pattern, correct, text, flags=re.IGNORECASE)

    applied_log.append("; ".join(summary) if summary else "")
    return text

# === Detect Anomalies ===
# Flags unknown words that:
    # Are not in glossary or ignore_words
    # Are not numbers or dice terms like 1d6
# Uses SpellChecker to find misspellings.
# =========================

def detect_anomalies(text, spell, glossary, ignore_words):
    words = re.findall(r"\b[\w']+\b", text)
    unknowns = []
    glossary_terms = set(k.lower() for k in glossary)
    known_misspellings = set(w.lower() for v in glossary.values() for w in v)

    for word in words:
        base = strip_suffix(word).lower()
        if base.isdigit() or re.fullmatch(r"\d*d\d+", base):
            continue
        if base in glossary_terms or base in known_misspellings or base in ignore_words:
            continue
        if not spell.known([base]):
            unknowns.append(base)
    return unknowns

# === Suggest New Glossary Entries ===
# Uses Levenshtein distance and Jaro-Winkler similarity to suggest replacements.
# Filters suggestions to keep only close matches.
# ====================================

def suggest_new_corrections(flagged_words, spell, glossary=None, max_suggestions=3):
    suggestions = {}
    glossary_terms = set(map(str.lower, (glossary or {}).keys()))

    for word in flagged_words:
        if word.lower() in glossary_terms:
            continue
        candidates = spell.candidates(word) or set()
        filtered = []
        for cand in candidates:
            lev_dist = Levenshtein.distance(word.lower(), cand.lower())
            jaro_sim = jellyfish.jaro_winkler_similarity(word.lower(), cand.lower())
            if lev_dist <= 2 and jaro_sim >= 0.85:
                filtered.append((cand, jaro_sim))
        filtered = sorted(filtered, key=lambda x: -x[1])[:max_suggestions]
        if filtered:
            suggestions[word] = [c for c, _ in filtered]
    return suggestions

# === Save Logs ===
# Creates a CSV showing:
    # Each anomaly
    # Count of appearances
    # Suggested fixes
# ==================

def save_anomaly_log(flagged_counts, suggestions):
    rows = []
    for word, count in flagged_counts.items():
        suggestion = ", ".join(suggestions.get(word, []))
        rows.append((word, count, suggestion))
    df = pd.DataFrame(rows, columns=["Word", "Count", "Suggested"])
    df.to_csv(ANOMALY_LOG_CSV, index=False)

# === Save Suggestions ===
# Flips wrong → [candidates] into candidate → [wrong words] so you can paste them into glossary_config.json.
# ========================

def save_suggestions_json(suggestions):
    flipped = defaultdict(list)
    for wrong, candidates in suggestions.items():
        for cand in candidates:
            flipped[cand].append(wrong)

    with open(SUGGESTIONS_FILE, "w", encoding="utf-8") as f:
        f.write("{\n")
        for i, (correct_term, wrong_list) in enumerate(sorted(flipped.items())):
            comma = "," if i < len(flipped) - 1 else ""
            json_list = json.dumps(sorted(set(wrong_list)))
            f.write(f'  "{correct_term}": {json_list}{comma}\n')
        f.write("}\n")

# === Print Suggestions ===
# Prints suggestions into the console - Uncomment snippet below if you want to enable this feature
# =========================

    # print("\n📘 Suggested additions to glossary_config.json:")
    # for correct_term, wrong_list in flipped.items():
        # print(f'  "{correct_term}": {json.dumps(sorted(set(wrong_list)))},')

# === Export .txt and .md Files ===
# Outputs a formatted transcript like:
    # [Traveller] I cast charm person.
    # [Jake] Okay, roll for it.
# Includes campaign header and a list of attendees.
# =================================

def export_txt_md(df, campaign, session_date, output_base):
    txt_path = output_base + ".txt"
    md_path = output_base + ".md"

    header = f"[{campaign}]\n[{session_date}]\nSession Attendance:\n"
    players = sorted(set(df["Player"]))

    with open(txt_path, "w", encoding="utf-8") as txt_file, open(md_path, "w", encoding="utf-8") as md_file:
        # Write headers
        txt_file.write(header)
        md_file.write(header)

        for player in players:
            txt_file.write(f"{player}\n")
            md_file.write(f"{player}\n")

        txt_file.write("\n")
        md_file.write("\n")

        # Write transcript in chat-style format
        for _, row in df.iterrows():
            line = f"[{row['Player']}] {row['Adjusted Line'].strip()}\n"
            txt_file.write(line)
            md_file.write(line)

# === Main Processing Function ===
# Full pipeline:
    # Load glossary
    # Combine lines
    # Apply corrections
    # Detect anomalies
    # Suggest glossary additions
    # Output .xlsx, .txt, .md, .csv, .json
# =================================

def process_file(file_path):
    print(f"📂 Processing: {file_path}")
    df = pd.read_excel(file_path)

    # Extract campaign name and session date from the filename
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    match = re.search(r'(\d{4}-\d{2}-\d{2})', base_name)
    if match:
        date = match.group(1)
        campaign = base_name[:match.start()].strip().replace("_", " ")
    else:
        campaign = base_name
        date = "Unknown"

    print(f"📘 Campaign: {campaign}")
    print(f"📅 Session Date: {date}")

    # === LOAD GLOSSARY (with campaign-specific logic) ===
    try:
        with open(GLOSSARY_FILE, "r", encoding="utf-8") as f:
            glossary_data = json.load(f)
    except Exception as e:
        print(f"❌ Failed to load {GLOSSARY_FILE}: {e}")
        return

    global_replace = glossary_data.get("Replace", {}).get("Global", {})
    campaign_replace = glossary_data.get("Replace", {}).get(campaign, {})
    ignore_words = set(word.lower() for word in glossary_data.get("Ignore", []))

    # Merge campaign-specific and global replacements
    replacements = global_replace.copy()
    for correct, wrongs in campaign_replace.items():
        if correct in replacements:
            replacements[correct] = sorted(set(replacements[correct] + wrongs))
        else:
            replacements[correct] = wrongs

    spell = SpellChecker()

    # === COMBINE LINES ===
    combined_df = combine_lines(df)

    # Assign consistent colors per player
    players = sorted(combined_df["Player"].unique())
    player_color_map = {
        player: SPEAKER_COLORS[i % len(SPEAKER_COLORS)]
        for i, player in enumerate(players)
    }

    # === APPLY CORRECTIONS ===
    correction_log = []
    combined_df["Adjusted Line"] = combined_df["Adjusted Line"].apply(
        lambda txt: apply_corrections(txt, replacements, correction_log)
    )
    combined_df["Corrections"] = correction_log

    # === ANOMALY DETECTION ===
    all_anomalies = []
    for text in combined_df["Adjusted Line"]:
        all_anomalies.extend(detect_anomalies(text, spell, replacements, ignore_words))

    anomaly_counts = Counter(all_anomalies)
    suggestions = suggest_new_corrections(anomaly_counts.keys(), spell, glossary=replacements)

    save_anomaly_log(anomaly_counts, suggestions)
    save_suggestions_json(suggestions)

    # === EXPORT ===
    output_base = os.path.splitext(file_path)[0] + "_corrected"
    xlsx_path = output_base + ".xlsx"
    combined_df[["Start", "End", "Player", "Adjusted Line", "Corrections"]].to_excel(xlsx_path, index=False)

    # Apply color highlights to Excel rows
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in range(2, ws.max_row + 1):
        player = ws[f"C{row}"].value
        color = player_color_map.get(player, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in "ABCDE":
            ws[f"{col}{row}"].fill = fill
    wb.save(xlsx_path)

    export_txt_md(combined_df, campaign, date, output_base)

    print(f"\n✅ Saved corrected transcript files:")
    print(f"  - {output_base}.xlsx")
    print(f"  - {output_base}.txt")
    print(f"  - {output_base}.md")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❗ Drag and drop a .xlsx transcript file onto this script.")
    else:
        process_file(sys.argv[1])